"""
Inventario Casa - Backend FastAPI Optimizado (Python Nativo)
Sistema de gestión de inventario, ventas, finanzas familiares y sincronización con Excel.

Características:
  - Backend dual: SQLite local (desarrollo) + Turso cloud (producción/Vercel)
  - Consultas instantáneas con Row dict-like + foreign keys
  - Tasa BCV en tiempo real con fallback inteligente
  - Mapeo exacto de reglas de negocio del Excel
  - Importación y exportación bidireccional en formato .xlsx (openpyxl)
"""

import os
import re
import time
import json
import sqlite3
import datetime
from contextlib import contextmanager
from typing import Optional, List, Dict, Any

import httpx
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Capa de Base de Datos Dual: SQLite local / Turso cloud (Vercel)
# ---------------------------------------------------------------------------

class _Row:
    """Wrapper que permite acceso por índice y por nombre de columna."""
    def __init__(self, columns, values):
        self._columns = columns
        self._values = values
    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        if isinstance(key, str):
            idx = self._columns.index(key)
            return self._values[idx]
        raise TypeError(f"Key must be int or str, got {type(key)}")
    def __contains__(self, key):
        return key in self._columns
    def keys(self):
        return self._columns
    def __iter__(self):
        return iter(self._columns)
    def __len__(self):
        return len(self._values)


class _TursoCursor:
    """Cursor-like wrapper para resultados Turso."""
    def __init__(self, columns, rows, lastrowid=None):
        self._columns = columns
        self._rows = rows
        self.lastrowid = lastrowid
        self.rowcount = len(rows)
    def fetchall(self):
        return [_Row(self._columns, r) for r in self._rows]
    def fetchone(self):
        return _Row(self._columns, self._rows[0]) if self._rows else None


def _turso_value(v):
    """Convierte un valor Python al formato de Value esperado por Turso pipeline."""
    if v is None:
        return "null"
    if isinstance(v, bool):
        return {"type": "integer", "value": "1" if v else "0"}
    if isinstance(v, int):
        return {"type": "integer", "value": str(v)}
    if isinstance(v, float):
        return {"type": "float", "value": float(v)}
    return {"type": "text", "value": str(v)}


def _turso_parse_rows(rows):
    """Convierte filas con formato Value de Turso a valores Python planos."""
    out = []
    for row in rows:
        parsed = []
        for cell in row:
            if isinstance(cell, dict):
                vt = cell.get("type")
                vv = cell.get("value")
                if vt == "integer":
                    parsed.append(int(vv))
                elif vt == "float":
                    parsed.append(float(vv))
                elif vt == "text":
                    parsed.append(vv)
                elif vt == "blob":
                    parsed.append(vv)
                elif vt == "null":
                    parsed.append(None)
                else:
                    parsed.append(vv)
            else:
                parsed.append(cell)
        out.append(parsed)
    return out


class _TursoConnection:
    """Conexi��n a Turso v��a REST API (httpx). Compatible con sqlite3 API."""
    def __init__(self, url, token):
        u = url.rstrip("/")
        if u.startswith("libsql://"):
            u = "https://" + u[len("libsql://"):]
        self._url = u
        self._token = token
    def execute(self, sql, params=()):
        args = [_turso_value(p) for p in (params or [])] if params else []
        resp = httpx.post(
            f"{self._url}/v2/pipeline",
            headers={
                "Authorization": f"Bearer {self._token}",
                "Content-Type": "application/json",
            },
            json={"requests": [{"type": "execute", "stmt": {"sql": sql, "args": args}}]},
            timeout=30.0,
        )
        resp.raise_for_status()
        data = resp.json()
        result = data["results"][0]
        if result.get("type") == "error":
            msg = result.get("error", {}).get("message", "") or result.get("message", "")
            if "UNIQUE" in msg or "constraint" in msg.lower():
                raise Exception(f"IntegrityError: {msg}")
            raise Exception(f"DB Error: {msg}")
        result_map = result.get("result") or result.get("response", {}).get("result", {})
        cols = [c["name"] for c in result_map.get("cols", [])]
        raw_rows = result_map.get("rows", [])
        rows = _turso_parse_rows(raw_rows)
        lastrowid = None
        lir = result_map.get("last_insert_rowid")
        if lir is not None:
            lastrowid = int(lir)
        return _TursoCursor(cols, rows, lastrowid=lastrowid)
    def executemany(self, sql, params_list):
        last = None
        count = 0
        for params in params_list:
            cur = self.execute(sql, params)
            last = cur
            count += cur.rowcount
        return last
    def commit(self):
        pass
    def close(self):
        pass


class _LocalConnection:
    """Envoltorio para sqlite3 local (desarrollo)."""
    def __init__(self, path):
        self._conn = sqlite3.connect(path)
        self._conn.row_factory = sqlite3.Row
        self._conn.execute("PRAGMA journal_mode=WAL")
        self._conn.execute("PRAGMA foreign_keys=ON")
    def execute(self, sql, params=()):
        return self._conn.execute(sql, params)
    def executemany(self, sql, params_list):
        return self._conn.executemany(sql, params_list)
    def commit(self):
        self._conn.commit()
    def rollback(self):
        self._conn.rollback()
    def close(self):
        self._conn.close()


TURSO_URL = os.environ.get("TURSO_DATABASE_URL", "")
TURSO_TOKEN = os.environ.get("TURSO_AUTH_TOKEN", "")


@contextmanager
def get_db():
    """Context manager para conexiones seguras. Turso en Vercel, SQLite local."""
    if TURSO_URL and TURSO_TOKEN:
        conn = _TursoConnection(TURSO_URL, TURSO_TOKEN)
    else:
        conn = _LocalConnection(DB_PATH)
    try:
        yield conn
        conn.commit()
    except Exception:
        if hasattr(conn, 'rollback'):
            conn.rollback()
        raise
    finally:
        conn.close()


def query_all(sql: str, params: tuple = ()) -> List[Dict[str, Any]]:
    """Ejecuta una consulta SELECT y retorna una lista de diccionarios."""
    with get_db() as conn:
        cur = conn.execute(sql, params)
        return [dict(zip(row._columns, row._values)) for row in cur.fetchall()]


def query_one(sql: str, params: tuple = ()) -> Optional[Dict[str, Any]]:
    """Ejecuta una consulta SELECT y retorna un diccionario o None."""
    with get_db() as conn:
        cur = conn.execute(sql, params)
        row = cur.fetchone()
        if row is None:
            return None
        return dict(zip(row._columns, row._values))


def execute_sql(sql: str, params: tuple = ()) -> Dict[str, Any]:
    """Ejecuta INSERT, UPDATE o DELETE y retorna metadatos."""
    with get_db() as conn:
        cur = conn.execute(sql, params)
        return {
            "last_insert_rowid": cur.lastrowid,
            "affected_rows": cur.rowcount,
        }

# ---------------------------------------------------------------------------
# Rutas y Constantes
# ---------------------------------------------------------------------------

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "inventario_casa.db")
PUBLIC_DIR = os.path.join(BASE_DIR, "public")
EXCEL_DESKTOP_PATH = os.path.join(
    os.path.expanduser("~"), "Desktop", "Inventario_Casa_Torres_Actualizado(2).xlsx"
)

app = FastAPI(
    title="Inventario Casa API",
    description="Sistema de gestión de inventario, ventas y finanzas familiares",
    version="2.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Manejo de Base de Datos — funciones auxiliares
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Esquema e Inicialización de Base de Datos
# ---------------------------------------------------------------------------

SCHEMA_SQL = [
    """
    CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL UNIQUE,
        categoria TEXT NOT NULL,
        precio_compra_actual REAL NOT NULL DEFAULT 0,
        precio_venta_actual REAL NOT NULL DEFAULT 0
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS movimientos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        tipo TEXT NOT NULL CHECK(tipo IN ('Compra', 'Venta')),
        producto_id INTEGER NOT NULL REFERENCES productos(id) ON DELETE RESTRICT,
        cantidad INTEGER NOT NULL CHECK(cantidad > 0),
        precio_unitario_bs REAL NOT NULL DEFAULT 0,
        costo_unitario_bs REAL NOT NULL DEFAULT 0,
        tasa_bcv REAL NOT NULL DEFAULT 0,
        persona_proveedor TEXT NOT NULL DEFAULT '',
        estado_pago TEXT NOT NULL CHECK(estado_pago IN ('Pagado', 'Fiado')),
        notas TEXT NOT NULL DEFAULT ''
    )
    """,
    """
    CREATE INDEX IF NOT EXISTS idx_movimientos_fecha ON movimientos(fecha);
    """,
    """
    CREATE INDEX IF NOT EXISTS idx_movimientos_tipo ON movimientos(tipo);
    """,
    """
    CREATE INDEX IF NOT EXISTS idx_movimientos_producto ON movimientos(producto_id);
    """
]

# Datos iniciales exactos del Excel para siembra si la BD está limpia
PRODUCTOS_INICIALES = [
    ("Cola Negra",  "Refresco", 840.0, 1300.0),
    ("Fresh",       "Refresco", 840.0, 1300.0),
    ("Uva",         "Refresco", 840.0, 1300.0),
    ("Colita",      "Refresco", 840.0, 1300.0),
    ("Polar Light", "Cerveza",  693.9175, 1100.0),
    ("Pilsen",      "Cerveza",  693.9175, 1100.0),
]


def init_database():
    """Inicializa tablas y siembra datos iniciales."""
    with get_db() as conn:
        for stmt in SCHEMA_SQL:
            conn.execute(stmt)

        # Verificar si ya hay datos (solo para Turso, en SQLite siempre funciona)
        if TURSO_URL and TURSO_TOKEN:
            try:
                cur = conn.execute("SELECT COUNT(*) FROM productos")
                row = cur.fetchone()
                prod_count = row[0] if row else 0
            except Exception:
                prod_count = 0
        else:
            cur = conn.execute("SELECT COUNT(*) FROM productos")
            prod_count = cur.fetchone()[0]

        if prod_count == 0:
            conn.executemany(
                "INSERT INTO productos (nombre, categoria, precio_compra_actual, precio_venta_actual) VALUES (?, ?, ?, ?)",
                PRODUCTOS_INICIALES,
            )
            print(f"[OK] {len(PRODUCTOS_INICIALES)} productos iniciales sembrados.")


@app.on_event("startup")
async def on_startup():
    init_database()
    print("[OK] Backend Inventario Casa iniciado exitosamente.")


# ---------------------------------------------------------------------------
# Lógica de Importación y Exportación con Excel
# ---------------------------------------------------------------------------

def normalizar_fecha(val: Any) -> str:
    """Convierte fechas de Excel o texto a formato YYYY-MM-DD."""
    if not val:
        return datetime.date.today().isoformat()
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    if "/" in val_str:
        parts = val_str.split("/")
        if len(parts) == 3:
            # DD/MM/YYYY
            return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
    return val_str


def importar_desde_excel_file(file_path_or_buffer, conn_externa=None):
    """Lee y carga el catálogo y todos los movimientos desde un archivo Excel."""
    wb = openpyxl.load_workbook(file_path_or_buffer, data_only=True)

    # 1. Encontrar hoja de catálogo
    ws_cat = None
    for name in wb.sheetnames:
        if "Cat" in name:
            ws_cat = wb[name]
            break

    # 2. Encontrar hoja de movimientos
    ws_mov = wb["Movimientos"] if "Movimientos" in wb.sheetnames else None

    # Tasa BCV del Dashboard si existe
    tasa_bcv_excel = 794.99
    if "Dashboard" in wb.sheetnames:
        ws_dash = wb["Dashboard"]
        for r in range(1, 10):
            c1 = ws_dash.cell(r, 1).value
            if c1 and "Tasa BCV" in str(c1):
                tasa_val = ws_dash.cell(r, 2).value
                if tasa_val and isinstance(tasa_val, (int, float)):
                    tasa_bcv_excel = float(tasa_val)
                break

    def _ejecutar(conn):
        # Insertar o actualizar productos
        prod_map = {}
        if ws_cat:
            for r in range(2, ws_cat.max_row + 1):
                pnom = ws_cat.cell(r, 2).value
                pcat = ws_cat.cell(r, 3).value
                pcomp = ws_cat.cell(r, 4).value or 0
                pvent = ws_cat.cell(r, 6).value or 0
                if pnom:
                    pnom_str = str(pnom).strip()
                    pcat_str = str(pcat).strip() if pcat else "General"
                    chk = conn.execute("SELECT id FROM productos WHERE nombre = ?", (pnom_str,)).fetchone()
                    if chk:
                        conn.execute(
                            "UPDATE productos SET categoria = ?, precio_compra_actual = ?, precio_venta_actual = ? WHERE id = ?",
                            (pcat_str, float(pcomp), float(pvent), chk["id"]),
                        )
                    else:
                        conn.execute(
                            "INSERT INTO productos (nombre, categoria, precio_compra_actual, precio_venta_actual) VALUES (?, ?, ?, ?)",
                            (pnom_str, pcat_str, float(pcomp), float(pvent)),
                        )

        # Mapear productos a IDs
        cur = conn.execute("SELECT id, nombre, precio_compra_actual, precio_venta_actual FROM productos")
        for row in cur.fetchall():
            prod_map[row["nombre"]] = dict(row)

        # Limpiar movimientos e insertar los del Excel
        if ws_mov:
            conn.execute("DELETE FROM movimientos")
            for r in range(2, ws_mov.max_row + 1):
                fecha_raw = ws_mov.cell(r, 1).value
                tipo = ws_mov.cell(r, 2).value
                prod_nom = ws_mov.cell(r, 3).value
                cant = ws_mov.cell(r, 4).value
                cat = ws_mov.cell(r, 5).value
                pu_bs = ws_mov.cell(r, 6).value
                persona = ws_mov.cell(r, 9).value
                pago = ws_mov.cell(r, 10).value
                notas = ws_mov.cell(r, 11).value

                if fecha_raw and tipo and prod_nom:
                    prod_nom_str = str(prod_nom).strip()
                    if prod_nom_str not in prod_map:
                        conn.execute(
                            "INSERT INTO productos (nombre, categoria, precio_compra_actual, precio_venta_actual) VALUES (?, ?, ?, ?)",
                            (prod_nom_str, str(cat or "General").strip(), float(pu_bs or 0), float(pu_bs or 0)),
                        )
                        cur_p = conn.execute("SELECT id, nombre, precio_compra_actual, precio_venta_actual FROM productos WHERE nombre = ?", (prod_nom_str,))
                        prod_map[prod_nom_str] = dict(cur_p.fetchone())

                    prod_info = prod_map[prod_nom_str]
                    fecha_norm = normalizar_fecha(fecha_raw)
                    cant_int = int(cant or 0)
                    tipo_str = str(tipo).strip()
                    pu_val = float(pu_bs or 0)

                    # Determinar costo y precio unitario
                    if tipo_str == "Compra":
                        costo_val = pu_val
                        precio_val = prod_info["precio_venta_actual"]
                    else:
                        costo_val = prod_info["precio_compra_actual"]
                        precio_val = pu_val

                    conn.execute(
                        """
                        INSERT INTO movimientos (
                            fecha, tipo, producto_id, cantidad,
                            precio_unitario_bs, costo_unitario_bs,
                            tasa_bcv, persona_proveedor, estado_pago, notas
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            fecha_norm,
                            tipo_str,
                            prod_info["id"],
                            cant_int,
                            precio_val,
                            costo_val,
                            tasa_bcv_excel,
                            str(persona or "").strip(),
                            str(pago or "Pagado").strip(),
                            str(notas or "").strip(),
                        ),
                    )

    if conn_externa:
        _ejecutar(conn_externa)
    else:
        with get_db() as conn:
            _ejecutar(conn)


def generar_libro_excel(tasa_bcv: float) -> openpyxl.Workbook:
    """Genera un archivo Excel profesional con las 4 hojas exactas del sistema."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # Estilos
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    section_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    bold_font = Font(name="Calibri", size=10, bold=True)
    regular_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # 1. Hoja Catálogo
    ws_cat = wb.create_sheet(title="Catálogo")
    cat_headers = ["ID", "Producto", "Categoría", "Precio Compra (Bs)", "Precio Compra (USD)", "Precio Venta (Bs)", "Precio Venta (USD)", "Stock"]
    ws_cat.append(cat_headers)
    for col_num in range(1, len(cat_headers) + 1):
        cell = ws_cat.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    productos = query_all("SELECT * FROM productos ORDER BY categoria, id")
    for row_idx, p in enumerate(productos, start=2):
        ws_cat.append([
            p["id"],
            p["nombre"],
            p["categoria"],
            p["precio_compra_actual"],
            f"=D{row_idx}/{tasa_bcv}",
            p["precio_venta_actual"],
            f"=F{row_idx}/{tasa_bcv}",
            f'=SUMIFS(Movimientos!D:D,Movimientos!C:C,B{row_idx},Movimientos!B:B,"Compra")-SUMIFS(Movimientos!D:D,Movimientos!C:C,B{row_idx},Movimientos!B:B,"Venta")',
        ])
        for c in range(1, 9):
            ws_cat.cell(row_idx, c).font = regular_font
            ws_cat.cell(row_idx, c).border = thin_border

    # 2. Hoja Movimientos
    ws_mov = wb.create_sheet(title="Movimientos")
    mov_headers = ["Fecha", "Tipo", "Producto", "Cantidad", "Categoría", "Precio Unitario (Bs)", "Total (Bs)", "Total (USD)", "Persona/Proveedor", "Estado Pago", "Notas"]
    ws_mov.append(mov_headers)
    for col_num in range(1, len(mov_headers) + 1):
        cell = ws_mov.cell(1, col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    movimientos = query_all("""
        SELECT m.*, p.nombre AS producto_nombre, p.categoria AS producto_categoria
        FROM movimientos m
        JOIN productos p ON m.producto_id = p.id
        ORDER BY m.fecha ASC, m.id ASC
    """)
    for row_idx, m in enumerate(movimientos, start=2):
        # Formato de fecha DD/MM/YYYY para Excel
        try:
            p_date = datetime.date.fromisoformat(m["fecha"])
            fecha_excel = p_date.strftime("%d/%m/%Y")
        except Exception:
            fecha_excel = m["fecha"]

        ws_mov.append([
            fecha_excel,
            m["tipo"],
            m["producto_nombre"],
            m["cantidad"],
            m["producto_categoria"],
            m["precio_unitario_bs"] if m["tipo"] == "Venta" else m["costo_unitario_bs"],
            f"=D{row_idx}*F{row_idx}",
            f"=G{row_idx}/{m['tasa_bcv'] if m['tasa_bcv'] > 0 else tasa_bcv}",
            m["persona_proveedor"],
            m["estado_pago"],
            m["notas"],
        ])
        for c in range(1, 12):
            ws_mov.cell(row_idx, c).font = regular_font
            ws_mov.cell(row_idx, c).border = thin_border

    # 3. Hoja Dashboard
    ws_dash = wb.create_sheet(title="Dashboard")
    ws_dash.cell(1, 1, "INVENTARIO CASA - DASHBOARD").font = title_font
    ws_dash.cell(3, 1, "Fecha Actual:").font = bold_font
    ws_dash.cell(3, 2, datetime.date.today().strftime("%d/%m/%Y")).font = regular_font
    ws_dash.cell(4, 1, "Tasa BCV (Bs/USD):").font = bold_font
    ws_dash.cell(4, 2, tasa_bcv).font = bold_font
    ws_dash.cell(5, 1, "Enlace BCV:").font = bold_font
    ws_dash.cell(5, 2, "https://www.bcv.org.ve/").font = regular_font

    # Secciones por categoría
    categorias = ["Refresco", "Cerveza"]
    cur_r = 7
    cat_ranges = {}

    for cat in categorias:
        plural = "REFRESCOS" if cat == "Refresco" else "CERVEZAS"
        ws_dash.cell(cur_r, 1, f"STOCK DE {plural}").font = section_font
        cur_r += 1
        ws_dash.cell(cur_r, 1, "Producto").font = bold_font
        ws_dash.cell(cur_r, 2, "Stock").font = bold_font
        ws_dash.cell(cur_r, 3, "Valor Stock (Bs)").font = bold_font
        ws_dash.cell(cur_r, 4, "Valor Stock (USD)").font = bold_font
        cur_r += 1

        start_cat_r = cur_r
        prods_cat = [p for p in productos if p["categoria"] == cat]
        for p in prods_cat:
            ws_dash.cell(cur_r, 1, p["nombre"]).font = regular_font
            ws_dash.cell(cur_r, 2, f'=SUMIFS(Movimientos!D:D,Movimientos!C:C,A{cur_r},Movimientos!B:B,"Compra")-SUMIFS(Movimientos!D:D,Movimientos!C:C,A{cur_r},Movimientos!B:B,"Venta")')
            ws_dash.cell(cur_r, 3, f'=B{cur_r}*IFERROR(VLOOKUP(A{cur_r},Catálogo!B:F,5,FALSE),0)')
            ws_dash.cell(cur_r, 4, f'=C{cur_r}/$B$4')
            cur_r += 1
        end_cat_r = cur_r - 1

        # Total categoría
        ws_dash.cell(cur_r, 1, f"TOTAL {plural}").font = bold_font
        ws_dash.cell(cur_r, 2, f"=SUM(B{start_cat_r}:B{end_cat_r})").font = bold_font
        ws_dash.cell(cur_r, 3, f"=SUM(C{start_cat_r}:C{end_cat_r})").font = bold_font
        ws_dash.cell(cur_r, 4, f"=SUM(D{start_cat_r}:D{end_cat_r})").font = bold_font
        cat_ranges[cat] = {"total_bs_cell": f"C{cur_r}", "prods": prods_cat}
        cur_r += 2

    # Resumen Financiero por Categoría
    resumen_cells = {}
    for cat in categorias:
        plural = "REFRESCOS" if cat == "Refresco" else "CERVEZAS"
        ws_dash.cell(cur_r, 1, f"RESUMEN FINANCIERO - {plural}").font = section_font
        cur_r += 1
        ws_dash.cell(cur_r, 1, "Concepto").font = bold_font
        ws_dash.cell(cur_r, 2, "Monto (Bs)").font = bold_font
        ws_dash.cell(cur_r, 3, "Monto (USD)").font = bold_font
        ws_dash.cell(cur_r, 4, "Detalle").font = bold_font
        cur_r += 1

        # Invertido
        r_inv = cur_r
        ws_dash.cell(cur_r, 1, "Total Invertido").font = regular_font
        ws_dash.cell(cur_r, 2, f'=SUMIFS(Movimientos!G:G,Movimientos!B:B,"Compra",Movimientos!E:E,"{cat}")')
        ws_dash.cell(cur_r, 3, f"=B{cur_r}/$B$4")
        ws_dash.cell(cur_r, 4, f"Compras de {cat.lower()}s").font = regular_font
        cur_r += 1

        # Fiado
        ws_dash.cell(cur_r, 1, "Total Fiado (Pendiente)").font = regular_font
        ws_dash.cell(cur_r, 2, f'=SUMIFS(Movimientos!G:G,Movimientos!B:B,"Venta",Movimientos!E:E,"{cat}",Movimientos!J:J,"Fiado")')
        ws_dash.cell(cur_r, 3, f"=B{cur_r}/$B$4")
        ws_dash.cell(cur_r, 4, f"Deuda clientes {cat.lower()}s").font = regular_font
        cur_r += 1

        # Vendido Pagado
        ws_dash.cell(cur_r, 1, "Total Vendido (Pagado)").font = regular_font
        ws_dash.cell(cur_r, 2, f'=SUMIFS(Movimientos!G:G,Movimientos!B:B,"Venta",Movimientos!E:E,"{cat}",Movimientos!J:J,"Pagado")')
        ws_dash.cell(cur_r, 3, f"=B{cur_r}/$B$4")
        ws_dash.cell(cur_r, 4, f"Ventas pagadas {cat.lower()}s").font = regular_font
        cur_r += 1

        # Vendido General
        ws_dash.cell(cur_r, 1, "Total Vendido (General)").font = regular_font
        ws_dash.cell(cur_r, 2, f'=SUMIFS(Movimientos!G:G,Movimientos!B:B,"Venta",Movimientos!E:E,"{cat}")')
        ws_dash.cell(cur_r, 3, f"=B{cur_r}/$B$4")
        ws_dash.cell(cur_r, 4, f"Todas las ventas {cat.lower()}s").font = regular_font
        cur_r += 1

        # Ganancia General
        # Sumar ganancias por cada producto
        prods_cat = cat_ranges[cat]["prods"]
        gan_parts = []
        for p in prods_cat:
            gan_parts.append(
                f'SUMIFS(Movimientos!G:G,Movimientos!B:B,"Venta",Movimientos!E:E,"{cat}",Movimientos!C:C,"{p["nombre"]}")-SUMIFS(Movimientos!D:D,Movimientos!B:B,"Venta",Movimientos!E:E,"{cat}",Movimientos!C:C,"{p["nombre"]}")*IFERROR(VLOOKUP("{p["nombre"]}",Catálogo!B:D,3,FALSE),0)'
            )
        gan_formula = "=" + "+".join(gan_parts) if gan_parts else "=0"

        r_gan = cur_r
        ws_dash.cell(cur_r, 1, "Ganancia General").font = bold_font
        ws_dash.cell(cur_r, 2, gan_formula).font = bold_font
        ws_dash.cell(cur_r, 3, f"=B{cur_r}/$B$4").font = bold_font
        ws_dash.cell(cur_r, 4, f"Ventas - Costo ({cat.lower()}s)").font = regular_font
        resumen_cells[f"gan_{cat}"] = f"B{cur_r}"
        cur_r += 1

        # Valor Stock Actual
        ws_dash.cell(cur_r, 1, "Valor Stock Actual").font = bold_font
        ws_dash.cell(cur_r, 2, f"={cat_ranges[cat]['total_bs_cell']}").font = bold_font
        ws_dash.cell(cur_r, 3, f"=B{cur_r}/$B$4").font = bold_font
        ws_dash.cell(cur_r, 4, f"Valor stock actual {cat.lower()}s").font = regular_font
        resumen_cells[f"stock_{cat}"] = f"B{cur_r}"
        cur_r += 2

    # Resumen General Consolidado
    ws_dash.cell(cur_r, 1, "RESUMEN GENERAL").font = section_font
    cur_r += 1
    ws_dash.cell(cur_r, 1, "Concepto").font = bold_font
    ws_dash.cell(cur_r, 2, "Monto (Bs)").font = bold_font
    ws_dash.cell(cur_r, 3, "Monto (USD)").font = bold_font
    ws_dash.cell(cur_r, 4, "Detalle").font = bold_font
    cur_r += 1

    # Ganancia Neta General
    ws_dash.cell(cur_r, 1, "GANANCIA NETA GENERAL").font = bold_font
    ws_dash.cell(cur_r, 2, f"={resumen_cells['gan_Refresco']}+{resumen_cells['gan_Cerveza']}").font = bold_font
    ws_dash.cell(cur_r, 3, f"=B{cur_r}/$B$4").font = bold_font
    ws_dash.cell(cur_r, 4, "Total ganado").font = regular_font
    cur_r += 1

    # Total General Valor Inventario
    ws_dash.cell(cur_r, 1, "TOTAL GENERAL (Valor Inventario)").font = bold_font
    ws_dash.cell(cur_r, 2, f"={resumen_cells['stock_Refresco']}+{resumen_cells['stock_Cerveza']}").font = bold_font
    ws_dash.cell(cur_r, 3, f"=B{cur_r}/$B$4").font = bold_font
    ws_dash.cell(cur_r, 4, "Valor total del inventario").font = regular_font

    # 4. Hoja Finanzas Mensuales
    ws_fin = wb.create_sheet(title="Finanzas")
    ws_fin.cell(1, 1, "CONTROL FINANCIERO MENSUAL").font = title_font

    cur_fin_r = 3
    meses_disponibles = ["2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]

    for cat in categorias:
        plural = "REFRESCOS" if cat == "Refresco" else "CERVEZAS"
        ws_fin.cell(cur_fin_r, 1, f"{plural} - MENSUAL").font = section_font
        cur_fin_r += 1
        fin_cols = ["Año-Mes", "Total Comprado (Bs)", "Total Vendido Pagado (Bs)", "Total Fiado (Bs)", "Total Vendido General (Bs)", "Costo Ventas (Bs)", "Ganancia (Bs)", "Inversión Acumulada (Bs)"]
        for i, h in enumerate(fin_cols, start=1):
            cell = ws_fin.cell(cur_fin_r, i, h)
            cell.font = bold_font
            cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        cur_fin_r += 1

        for mes in meses_disponibles:
            ws_fin.cell(cur_fin_r, 1, mes).font = regular_font
            ws_fin.cell(cur_fin_r, 2, f'=SUMIFS(Movimientos!G:G,Movimientos!B:B,"Compra",Movimientos!E:E,"{cat}")')
            ws_fin.cell(cur_fin_r, 3, f'=SUMIFS(Movimientos!G:G,Movimientos!B:B,"Venta",Movimientos!E:E,"{cat}",Movimientos!J:J,"Pagado")')
            ws_fin.cell(cur_fin_r, 4, f'=SUMIFS(Movimientos!G:G,Movimientos!B:B,"Venta",Movimientos!E:E,"{cat}",Movimientos!J:J,"Fiado")')
            ws_fin.cell(cur_fin_r, 5, f'=SUMIFS(Movimientos!G:G,Movimientos!B:B,"Venta",Movimientos!E:E,"{cat}")')
            ws_fin.cell(cur_fin_r, 6, f'=E{cur_fin_r}-G{cur_fin_r}')
            ws_fin.cell(cur_fin_r, 7, f'=E{cur_fin_r}-F{cur_fin_r}')
            ws_fin.cell(cur_fin_r, 8, f'=B{cur_fin_r}-E{cur_fin_r}')
            cur_fin_r += 1
        cur_fin_r += 2

    # Ajustar anchos de columnas
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len and not val_str.startswith("="):
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    return wb


# ---------------------------------------------------------------------------
# Modelos Pydantic para la API
# ---------------------------------------------------------------------------

class ProductoCreate(BaseModel):
    nombre: str
    categoria: str
    precio_compra_actual: float
    precio_venta_actual: float


class ProductoUpdate(BaseModel):
    precio_compra_actual: float
    precio_venta_actual: float
    categoria: Optional[str] = None


class MovimientoCreate(BaseModel):
    fecha: Optional[str] = None
    tipo: str
    producto_id: int
    cantidad: int
    tasa_bcv: Optional[float] = None
    precio_unitario_bs: Optional[float] = None
    persona_proveedor: Optional[str] = ""
    estado_pago: str
    notas: Optional[str] = ""


class MovimientoUpdate(BaseModel):
    fecha: str
    tipo: str
    producto_id: int
    cantidad: int
    precio_unitario_bs: float
    costo_unitario_bs: float
    tasa_bcv: float
    persona_proveedor: str
    estado_pago: str
    notas: str


# ---------------------------------------------------------------------------
# Endpoints de la API
# ---------------------------------------------------------------------------

# Cache en memoria para evitar saturar el portal del BCV en cada petición
_BCV_CACHE = {
    "data": None,
    "timestamp": 0.0,
}

BCV_URL = "https://www.bcv.org.ve/"
BCV_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
}


def parse_bcv_html(html: str) -> Optional[Dict[str, Any]]:
    """
    Extrae la tasa oficial del dólar y fecha valor del HTML del portal oficial del BCV.
    Maneja el formato numérico venezolano (coma como separador decimal).
    """
    idx = html.find('id="dolar"')
    if idx == -1:
        idx = html.find('id="view-dolar"')
    if idx == -1:
        idx = html.find("dolar")

    if idx == -1:
        return None

    snippet = html[idx : idx + 800]
    match = re.search(r"<strong[^>]*>\s*([0-9.,]+)\s*</strong>", snippet, re.IGNORECASE)
    if not match:
        return None

    raw_val = match.group(1).strip()
    if "," in raw_val:
        clean_val = raw_val.replace(".", "").replace(",", ".")
    else:
        clean_val = raw_val

    try:
        tasa = round(float(clean_val), 4)
    except ValueError:
        return None

    fecha_valor = None
    m_fecha = re.search(r'class="date-display-single"[^>]*content="([^"]+)"', html, re.IGNORECASE)
    if not m_fecha:
        m_fecha = re.search(r'<span class="date-display-single"[^>]*>([^<]+)</span>', html, re.IGNORECASE)
    if m_fecha:
        fecha_valor = m_fecha.group(1).strip()

    return {
        "tasa": tasa,
        "fecha_actualizacion": fecha_valor or datetime.datetime.now().isoformat(),
        "nombre": "Dólar BCV Oficial",
        "fuente": "Banco Central de Venezuela (bcv.org.ve)",
    }


async def scrape_bcv_directo() -> Optional[Dict[str, Any]]:
    """Realiza la petición directa al portal oficial del BCV (https://www.bcv.org.ve/)."""
    try:
        async with httpx.AsyncClient(verify=False, timeout=12.0, headers=BCV_HEADERS, follow_redirects=True) as client:
            resp = await client.get(BCV_URL)
            if resp.status_code == 200:
                data = parse_bcv_html(resp.text)
                if data and data["tasa"] > 0:
                    return data
    except Exception as e:
        print(f"[Aviso Scraping BCV]: No se pudo conectar directamente con bcv.org.ve: {e}")
    return None


@app.get("/api/bcv")
async def get_bcv(refresh: bool = False):
    """
    Consulta la tasa oficial del BCV en tiempo real.
    1. Intenta scraping directo del portal oficial bcv.org.ve
    2. Fallback secundario a ve.dolarapi.com
    3. Fallback terciario a la última tasa registrada en la BD local
    """
    ahora = time.time()
    if not refresh and _BCV_CACHE["data"] and (ahora - _BCV_CACHE["timestamp"] < 600):
        return _BCV_CACHE["data"]

    # 1. Scraping Directo Oficial
    data_oficial = await scrape_bcv_directo()
    if data_oficial:
        _BCV_CACHE["data"] = data_oficial
        _BCV_CACHE["timestamp"] = ahora
        return data_oficial

    # 2. Fallback Secundario (API Espejo)
    try:
        async with httpx.AsyncClient(timeout=6.0) as client:
            resp = await client.get("https://ve.dolarapi.com/v1/dolares/oficial")
            if resp.status_code == 200:
                data = resp.json()
                tasa = float(data.get("promedio", data.get("precio", 0)))
                res = {
                    "tasa": tasa,
                    "fecha_actualizacion": data.get("fechaActualizacion", datetime.datetime.now().isoformat()),
                    "nombre": data.get("nombre", "Dólar BCV (Espejo)"),
                    "fuente": "ve.dolarapi.com (Oficial BCV)",
                }
                _BCV_CACHE["data"] = res
                _BCV_CACHE["timestamp"] = ahora
                return res
    except Exception:
        pass

    # 3. Fallback Terciario (Última tasa histórica en SQLite)
    row = query_one("SELECT tasa_bcv FROM movimientos WHERE tasa_bcv > 0 ORDER BY id DESC LIMIT 1")
    tasa_fallback = row["tasa_bcv"] if row and row["tasa_bcv"] else 804.81

    return {
        "tasa": tasa_fallback,
        "fecha_actualizacion": datetime.datetime.now().isoformat(),
        "nombre": "Dólar BCV (Último registrado)",
        "fuente": "Historial Local",
    }


@app.get("/api/productos")
async def get_productos():
    """Retorna el catálogo completo con stock calculado en tiempo real."""
    sql = """
        SELECT
            p.id,
            p.nombre,
            p.categoria,
            p.precio_compra_actual,
            p.precio_venta_actual,
            COALESCE(SUM(CASE WHEN m.tipo = 'Compra' THEN m.cantidad ELSE 0 END), 0)
            - COALESCE(SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad ELSE 0 END), 0) AS stock
        FROM productos p
        LEFT JOIN movimientos m ON p.id = m.producto_id
        GROUP BY p.id
        ORDER BY p.categoria, p.nombre
    """
    return query_all(sql)


@app.post("/api/productos", status_code=201)
async def create_producto(data: ProductoCreate):
    """Crea un nuevo producto en el catálogo."""
    try:
        res = execute_sql(
            "INSERT INTO productos (nombre, categoria, precio_compra_actual, precio_venta_actual) VALUES (?, ?, ?, ?)",
            (data.nombre.strip(), data.categoria.strip(), data.precio_compra_actual, data.precio_venta_actual),
        )
        return {"id": res["last_insert_rowid"], "mensaje": "Producto creado con éxito"}
    except Exception as e:
        if "UNIQUE" in str(e) or "constraint" in str(e).lower():
            raise HTTPException(400, detail="Ya existe un producto con este nombre")
        raise HTTPException(500, detail=f"Error creando producto: {str(e)}")


@app.put("/api/productos/{producto_id}")
async def update_producto(producto_id: int, data: ProductoUpdate):
    """Actualiza los precios y categoría de un producto."""
    prod = query_one("SELECT * FROM productos WHERE id = ?", (producto_id,))
    if not prod:
        raise HTTPException(404, detail="Producto no encontrado")

    cat = data.categoria.strip() if data.categoria else prod["categoria"]
    execute_sql(
        "UPDATE productos SET precio_compra_actual = ?, precio_venta_actual = ?, categoria = ? WHERE id = ?",
        (data.precio_compra_actual, data.precio_venta_actual, cat, producto_id),
    )
    return {"mensaje": "Producto actualizado correctamente", "id": producto_id}


@app.delete("/api/productos/{producto_id}")
async def delete_producto(producto_id: int):
    """Elimina un producto si no tiene movimientos registrados."""
    movs = query_one("SELECT COUNT(*) AS cnt FROM movimientos WHERE producto_id = ?", (producto_id,))
    if movs and movs["cnt"] > 0:
        raise HTTPException(400, detail=f"No se puede eliminar: tiene {movs['cnt']} movimientos registrados")

    execute_sql("DELETE FROM productos WHERE id = ?", (producto_id,))
    return {"mensaje": "Producto eliminado"}


@app.get("/api/movimientos")
async def get_movimientos():
    """Retorna el historial completo de movimientos con cálculos en Bs y USD."""
    sql = """
        SELECT
            m.id,
            m.fecha,
            m.tipo,
            m.producto_id,
            p.nombre AS producto_nombre,
            p.categoria AS producto_categoria,
            m.cantidad,
            m.precio_unitario_bs,
            m.costo_unitario_bs,
            m.tasa_bcv,
            m.persona_proveedor,
            m.estado_pago,
            m.notas,
            ROUND(m.cantidad * (CASE WHEN m.tipo = 'Compra' THEN m.costo_unitario_bs ELSE m.precio_unitario_bs END), 2) AS total_bs,
            ROUND((m.cantidad * (CASE WHEN m.tipo = 'Compra' THEN m.costo_unitario_bs ELSE m.precio_unitario_bs END)) / CASE WHEN m.tasa_bcv > 0 THEN m.tasa_bcv ELSE 1 END, 2) AS total_usd,
            ROUND(CASE WHEN m.tipo = 'Venta' THEN m.cantidad * (m.precio_unitario_bs - m.costo_unitario_bs) ELSE 0 END, 2) AS ganancia_bs
        FROM movimientos m
        JOIN productos p ON m.producto_id = p.id
        ORDER BY m.fecha DESC, m.id DESC
    """
    return query_all(sql)


@app.post("/api/movimientos", status_code=201)
async def create_movimiento(data: MovimientoCreate):
    """
    Registra un movimiento congelando los precios históricos vigentes.
    Si es compra, permite especificar el costo unitario de compra directamente.
    """
    prod = query_one("SELECT * FROM productos WHERE id = ?", (data.producto_id,))
    if not prod:
        raise HTTPException(404, detail="Producto no encontrado")

    # Obtener tasa BCV si no viene provista
    tasa = data.tasa_bcv
    if not tasa or tasa <= 0:
        bcv_info = await get_bcv()
        tasa = bcv_info["tasa"]

    fecha_reg = data.fecha or datetime.date.today().isoformat()

    if data.tipo == "Compra":
        costo_unitario = data.precio_unitario_bs if data.precio_unitario_bs is not None and data.precio_unitario_bs > 0 else prod["precio_compra_actual"]
        precio_unitario = prod["precio_venta_actual"]
    else:
        # Venta
        costo_unitario = prod["precio_compra_actual"]
        precio_unitario = data.precio_unitario_bs if data.precio_unitario_bs is not None and data.precio_unitario_bs > 0 else prod["precio_venta_actual"]

    res = execute_sql(
        """
        INSERT INTO movimientos (
            fecha, tipo, producto_id, cantidad,
            precio_unitario_bs, costo_unitario_bs,
            tasa_bcv, persona_proveedor, estado_pago, notas
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            fecha_reg,
            data.tipo,
            data.producto_id,
            data.cantidad,
            precio_unitario,
            costo_unitario,
            tasa,
            (data.persona_proveedor or "").strip(),
            data.estado_pago,
            (data.notas or "").strip(),
        ),
    )

    # Si es una Compra, actualizar el precio_compra_actual del producto
    # para que el catálogo refleje siempre el último precio pagado al proveedor
    if data.tipo == "Compra":
        execute_sql(
            "UPDATE productos SET precio_compra_actual = ? WHERE id = ?",
            (costo_unitario, data.producto_id),
        )

    return {
        "id": res["last_insert_rowid"],
        "mensaje": "Movimiento registrado correctamente",
        "precio_unitario_bs": precio_unitario,
        "costo_unitario_bs": costo_unitario,
    }


@app.put("/api/movimientos/{movimiento_id}")
async def update_movimiento(movimiento_id: int, data: MovimientoUpdate):
    """Modifica un movimiento existente."""
    mov = query_one("SELECT id FROM movimientos WHERE id = ?", (movimiento_id,))
    if not mov:
        raise HTTPException(404, detail="Movimiento no encontrado")

    execute_sql(
        """
        UPDATE movimientos SET
            fecha = ?, tipo = ?, producto_id = ?, cantidad = ?,
            precio_unitario_bs = ?, costo_unitario_bs = ?,
            tasa_bcv = ?, persona_proveedor = ?, estado_pago = ?, notas = ?
        WHERE id = ?
        """,
        (
            data.fecha, data.tipo, data.producto_id, data.cantidad,
            data.precio_unitario_bs, data.costo_unitario_bs,
            data.tasa_bcv, data.persona_proveedor.strip(), data.estado_pago, data.notas.strip(),
            movimiento_id
        ),
    )
    return {"mensaje": "Movimiento actualizado", "id": movimiento_id}


@app.delete("/api/movimientos/{movimiento_id}")
async def delete_movimiento(movimiento_id: int):
    """Elimina un movimiento registrado."""
    execute_sql("DELETE FROM movimientos WHERE id = ?", (movimiento_id,))
    return {"mensaje": "Movimiento eliminado"}


@app.put("/api/movimientos/{movimiento_id}/pagar")
async def marcar_como_pagado(movimiento_id: int):
    """Marca una venta fiada como 'Pagado'."""
    mov = query_one("SELECT * FROM movimientos WHERE id = ?", (movimiento_id,))
    if not mov:
        raise HTTPException(404, detail="Movimiento no encontrado")

    execute_sql("UPDATE movimientos SET estado_pago = 'Pagado' WHERE id = ?", (movimiento_id,))
    return {"mensaje": "Deuda saldada correctamente", "id": movimiento_id}


@app.get("/api/fiados")
async def get_fiados():
    """
    Retorna el estado de cuentas por cobrar agrupado por persona/cliente,
    con detalle de cada venta pendiente y montos totales en Bs y USD.
    """
    sql = """
        SELECT
            m.id,
            m.fecha,
            m.persona_proveedor AS cliente,
            p.nombre AS producto_nombre,
            p.categoria AS producto_categoria,
            m.cantidad,
            m.precio_unitario_bs,
            m.tasa_bcv,
            m.notas,
            ROUND(m.cantidad * m.precio_unitario_bs, 2) AS total_bs,
            ROUND((m.cantidad * m.precio_unitario_bs) / CASE WHEN m.tasa_bcv > 0 THEN m.tasa_bcv ELSE 1 END, 2) AS total_usd
        FROM movimientos m
        JOIN productos p ON m.producto_id = p.id
        WHERE m.tipo = 'Venta' AND m.estado_pago = 'Fiado'
        ORDER BY m.persona_proveedor, m.fecha DESC
    """
    rows = query_all(sql)

    # Agrupar por persona
    clientes: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        nom = r["cliente"] or "Sin nombre"
        if nom not in clientes:
            clientes[nom] = {
                "cliente": nom,
                "total_deuda_bs": 0.0,
                "total_deuda_usd": 0.0,
                "cantidad_items": 0,
                "movimientos": [],
            }
        clientes[nom]["total_deuda_bs"] += r["total_bs"]
        clientes[nom]["total_deuda_usd"] += r["total_usd"]
        clientes[nom]["cantidad_items"] += r["cantidad"]
        clientes[nom]["movimientos"].append(r)

    lista_clientes = sorted(clientes.values(), key=lambda x: x["total_deuda_bs"], reverse=True)
    total_general_bs = sum(c["total_deuda_bs"] for c in lista_clientes)
    total_general_usd = sum(c["total_deuda_usd"] for c in lista_clientes)

    return {
        "total_general_bs": round(total_general_bs, 2),
        "total_general_usd": round(total_general_usd, 2),
        "total_deudores": len(lista_clientes),
        "clientes": lista_clientes,
    }


@app.get("/api/dashboard")
async def get_dashboard():
    """
    Retorna todas las métricas del sistema calculadas exactamente como en la hoja Dashboard del Excel:
    - Tasa BCV
    - Stock por producto y por categoría (Refrescos, Cervezas)
    - Resumen financiero por categoría (Invertido, Fiado, Vendido Pagado, Vendido General, Ganancia General, Valor Stock)
    - Resumen General Consolidado (Ganancia Neta General, Total General Valor Inventario)
    """
    # 1. Tasa BCV actual
    bcv_res = await get_bcv()
    tasa_bcv = bcv_res["tasa"]

    # 2. Stock y valorización por producto
    sql_stock = """
        SELECT
            p.id,
            p.nombre,
            p.categoria,
            p.precio_compra_actual,
            p.precio_venta_actual,
            COALESCE(SUM(CASE WHEN m.tipo = 'Compra' THEN m.cantidad ELSE 0 END), 0)
            - COALESCE(SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad ELSE 0 END), 0) AS stock
        FROM productos p
        LEFT JOIN movimientos m ON p.id = m.producto_id
        GROUP BY p.id
        ORDER BY p.categoria, p.nombre
    """
    stock_rows = query_all(sql_stock)

    stock_por_categoria: Dict[str, Dict[str, Any]] = {}
    for row in stock_rows:
        cat = row["categoria"]
        if cat not in stock_por_categoria:
            stock_por_categoria[cat] = {
                "categoria": cat,
                "unidades_totales": 0,
                "valor_stock_bs": 0.0,
                "valor_stock_usd": 0.0,
                "productos": [],
            }
        unidades = row["stock"]
        valor_bs = round(unidades * row["precio_venta_actual"], 2)
        valor_usd = round(valor_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0

        item = {
            "id": row["id"],
            "nombre": row["nombre"],
            "stock": unidades,
            "precio_venta_bs": row["precio_venta_actual"],
            "precio_venta_usd": round(row["precio_venta_actual"] / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "precio_compra_bs": row["precio_compra_actual"],
            "precio_compra_usd": round(row["precio_compra_actual"] / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "valor_stock_bs": valor_bs,
            "valor_stock_usd": valor_usd,
        }
        stock_por_categoria[cat]["productos"].append(item)
        stock_por_categoria[cat]["unidades_totales"] += unidades
        stock_por_categoria[cat]["valor_stock_bs"] += valor_bs
        stock_por_categoria[cat]["valor_stock_usd"] += valor_usd

    # 3. Métricas financieras por categoría
    sql_fin_cat = """
        SELECT
            p.categoria,
            ROUND(SUM(CASE WHEN m.tipo = 'Compra' THEN m.cantidad * m.costo_unitario_bs ELSE 0 END), 2) AS total_invertido_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' AND m.estado_pago = 'Fiado' THEN m.cantidad * m.precio_unitario_bs ELSE 0 END), 2) AS total_fiado_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' AND m.estado_pago = 'Pagado' THEN m.cantidad * m.precio_unitario_bs ELSE 0 END), 2) AS total_vendido_pagado_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad * m.precio_unitario_bs ELSE 0 END), 2) AS total_vendido_general_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad * m.costo_unitario_bs ELSE 0 END), 2) AS costo_ventas_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad * (m.precio_unitario_bs - m.costo_unitario_bs) ELSE 0 END), 2) AS ganancia_general_bs
        FROM productos p
        LEFT JOIN movimientos m ON p.id = m.producto_id
        GROUP BY p.categoria
    """
    fin_cat_rows = query_all(sql_fin_cat)

    resumen_financiero_cat: Dict[str, Dict[str, Any]] = {}
    for r in fin_cat_rows:
        cat = r["categoria"]
        inv_bs = r["total_invertido_bs"] or 0.0
        fiado_bs = r["total_fiado_bs"] or 0.0
        pagado_bs = r["total_vendido_pagado_bs"] or 0.0
        vendido_bs = r["total_vendido_general_bs"] or 0.0
        costo_bs = r["costo_ventas_bs"] or 0.0
        gan_bs = r["ganancia_general_bs"] or 0.0
        stock_val_bs = stock_por_categoria.get(cat, {}).get("valor_stock_bs", 0.0)

        resumen_financiero_cat[cat] = {
            "categoria": cat,
            "total_invertido_bs": inv_bs,
            "total_invertido_usd": round(inv_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "total_fiado_bs": fiado_bs,
            "total_fiado_usd": round(fiado_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "total_vendido_pagado_bs": pagado_bs,
            "total_vendido_pagado_usd": round(pagado_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "total_vendido_general_bs": vendido_bs,
            "total_vendido_general_usd": round(vendido_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "costo_ventas_bs": costo_bs,
            "costo_ventas_usd": round(costo_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "ganancia_general_bs": gan_bs,
            "ganancia_general_usd": round(gan_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "valor_stock_actual_bs": stock_val_bs,
            "valor_stock_actual_usd": round(stock_val_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
        }

    # 4. Totales Globales
    total_inventario_bs = sum(c["valor_stock_bs"] for c in stock_por_categoria.values())
    total_ganancia_bs = sum(c["ganancia_general_bs"] for c in resumen_financiero_cat.values())
    total_fiado_global_bs = sum(c["total_fiado_bs"] for c in resumen_financiero_cat.values())
    total_invertido_global_bs = sum(c["total_invertido_bs"] for c in resumen_financiero_cat.values())
    total_vendido_global_bs = sum(c["total_vendido_general_bs"] for c in resumen_financiero_cat.values())

    return {
        "tasa_bcv": tasa_bcv,
        "resumen_general": {
            "ganancia_neta_bs": round(total_ganancia_bs, 2),
            "ganancia_neta_usd": round(total_ganancia_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "valor_inventario_bs": round(total_inventario_bs, 2),
            "valor_inventario_usd": round(total_inventario_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "total_fiado_bs": round(total_fiado_global_bs, 2),
            "total_fiado_usd": round(total_fiado_global_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "total_invertido_bs": round(total_invertido_global_bs, 2),
            "total_invertido_usd": round(total_invertido_global_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
            "total_vendido_bs": round(total_vendido_global_bs, 2),
            "total_vendido_usd": round(total_vendido_global_bs / tasa_bcv, 2) if tasa_bcv > 0 else 0.0,
        },
        "stock_por_categoria": stock_por_categoria,
        "resumen_por_categoria": resumen_financiero_cat,
    }


@app.get("/api/finanzas")
async def get_finanzas():
    """Retorna el control financiero mensual agrupado por mes y por categoría."""
    bcv_res = await get_bcv()
    tasa_bcv = bcv_res["tasa"]

    sql = """
        SELECT
            STRFTIME('%Y-%m', m.fecha) AS mes,
            p.categoria,
            ROUND(SUM(CASE WHEN m.tipo = 'Compra' THEN m.cantidad * m.costo_unitario_bs ELSE 0 END), 2) AS total_comprado_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' AND m.estado_pago = 'Pagado' THEN m.cantidad * m.precio_unitario_bs ELSE 0 END), 2) AS total_vendido_pagado_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' AND m.estado_pago = 'Fiado' THEN m.cantidad * m.precio_unitario_bs ELSE 0 END), 2) AS total_fiado_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad * m.precio_unitario_bs ELSE 0 END), 2) AS total_vendido_general_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad * m.costo_unitario_bs ELSE 0 END), 2) AS costo_ventas_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad * (m.precio_unitario_bs - m.costo_unitario_bs) ELSE 0 END), 2) AS ganancia_bs,
            ROUND(SUM(CASE WHEN m.tipo = 'Compra' THEN m.cantidad * m.costo_unitario_bs ELSE 0 END) - SUM(CASE WHEN m.tipo = 'Venta' THEN m.cantidad * m.precio_unitario_bs ELSE 0 END), 2) AS inversion_acumulada_bs
        FROM movimientos m
        JOIN productos p ON m.producto_id = p.id
        GROUP BY mes, p.categoria
        ORDER BY mes DESC, p.categoria
    """
    rows = query_all(sql)

    resultado = []
    for r in rows:
        item = dict(r)
        item["total_comprado_usd"] = round((item["total_comprado_bs"] or 0) / tasa_bcv, 2) if tasa_bcv > 0 else 0.0
        item["total_vendido_pagado_usd"] = round((item["total_vendido_pagado_bs"] or 0) / tasa_bcv, 2) if tasa_bcv > 0 else 0.0
        item["total_fiado_usd"] = round((item["total_fiado_bs"] or 0) / tasa_bcv, 2) if tasa_bcv > 0 else 0.0
        item["total_vendido_general_usd"] = round((item["total_vendido_general_bs"] or 0) / tasa_bcv, 2) if tasa_bcv > 0 else 0.0
        item["ganancia_usd"] = round((item["ganancia_bs"] or 0) / tasa_bcv, 2) if tasa_bcv > 0 else 0.0
        resultado.append(item)

    return resultado


@app.post("/api/excel/importar")
async def importar_excel(file: Optional[UploadFile] = File(None)):
    """Importa el catálogo y movimientos desde un archivo subido o directo del escritorio."""
    try:
        if file:
            importar_desde_excel_file(file.file)
            origen = f"archivo subido ({file.filename})"
        else:
            if not os.path.exists(EXCEL_DESKTOP_PATH):
                raise HTTPException(404, detail=f"No se encontró el archivo en {EXCEL_DESKTOP_PATH}")
            importar_desde_excel_file(EXCEL_DESKTOP_PATH)
            origen = f"escritorio ({EXCEL_DESKTOP_PATH})"

        prod_count = query_one("SELECT COUNT(*) AS cnt FROM productos")["cnt"]
        mov_count = query_one("SELECT COUNT(*) AS cnt FROM movimientos")["cnt"]

        return {
            "mensaje": f"Sincronización completada exitosamente desde {origen}",
            "productos_totales": prod_count,
            "movimientos_totales": mov_count,
        }
    except Exception as e:
        raise HTTPException(500, detail=f"Error importando Excel: {str(e)}")


@app.get("/api/excel/exportar")
async def exportar_excel():
    """Genera y descarga el archivo Excel con las 4 hojas exactas del sistema."""
    bcv_res = await get_bcv()
    tasa_bcv = bcv_res["tasa"]

    wb = generar_libro_excel(tasa_bcv)

    temp_path = os.path.join(BASE_DIR, "temp_inventario.xlsx")
    wb.save(temp_path)

    filename = f"Inventario_Casa_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    return FileResponse(
        temp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


# ---------------------------------------------------------------------------
# Frontend SPA Estático
# ---------------------------------------------------------------------------

if os.path.exists(PUBLIC_DIR):
    app.mount("/public", StaticFiles(directory=PUBLIC_DIR), name="public")

    @app.get("/")
    async def serve_spa():
        index_file = os.path.join(PUBLIC_DIR, "index.html")
        if os.path.exists(index_file):
            return FileResponse(index_file)
        return {"mensaje": "Inventario Casa API Activa"}
